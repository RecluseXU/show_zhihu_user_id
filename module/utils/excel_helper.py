#!/usr/bin/env python
# -*- encoding: utf-8 -*-
'''
@File    :   baidu_check.py
@Time    :   2020.9.9
@Author  :   EvilRecluse
@Version :   1.0
@Contact :   evilrecluse@sxkid.com
@Doc     :
    操作 excel 数据的工具
    有两个库可以用  openpyxl 和 xlrd
        xlrd 
            不能执行写操作，但读取速度较快，垃圾会自动回收
        openpyxl 
            相对xlrd来说，这个库比较新，能执行写操作
            但处理速度较慢，没有方便的选取非空行的能力
            除非开启只读模式，要不然无法手动回收垃圾
'''

from openpyxl import load_workbook
import xlrd


class PyExcel(object):
    @staticmethod
    def get_data_simple(file_path: str, sheet_name: str, data_format: dict, rowlimit=[0, "max"]) -> list:
        '''
        静态函数版本，从给定的excel文件中的给定的sheet中获取数据
        若是有某行一条所需数据记录都没有，那么会跳过此行
        params:
            file_path   excel文件路径
            sheet_name  目标Sheet名字
            data_format 数据获取字典，将根据字典设置返回结果
                {数据名: 列号(从0开始), ......}
            rowlimit    行号限制，只返回限制行内的数据
                [int, int]
        '''
        book = xlrd.open_workbook(file_path)
        sheet = book.sheet_by_name(sheet_name)

        rowlimit[1] = sheet.nrows if rowlimit[1] == "max" else rowlimit[1]

        data_column = data_format.items()

        _result = []
        for i in range(rowlimit[0], rowlimit[1]):
            note = {
                _key: sheet.cell(i, _val).value
                for _key, _val in data_column
            }
            # 如果记录中全是空的
            if sum(map(lambda x: x!="", note.values())) == 0:
                continue
            _result. append(note)
        return _result


        # return [
        #     {
        #         _key: sheet.cell(i, _val).value
        #         for _key, _val in data_column
        #     }
        #     for i in range(rowlimit[0], rowlimit[1])]


    @staticmethod
    def write_data(input_file_path: str, output_file_path: str, sheet_name: str, primary_key_col: int, data: dict, insert_col_list=[]):
        '''
        写数据到excel中
        input_file_path  原excel文件地址
        output_file_path 输出的excel文件地址
        primary_key_col  主键列数，用于寻找位置
        data_format  数据填写字典，将根据字典填入数据
            {主键值: {单元格列数: 单项数据......}}
        insert_col       可以在列表中填入数据可以添加列，添加列的操作会在写入操作之前执行
        '''
        wb = load_workbook(input_file_path)
        worksheet = wb[sheet_name]

        # 插入列
        for i in insert_col_list:
            worksheet.work_sheet.insert_cols(i)
        # 写数据
        for i in range(1, worksheet.max_row):
            primary_key_value = worksheet.cell(i, primary_key_col).value
            print(primary_key_value)
            if primary_key_value and primary_key_value in data.keys():
                for col_ind, val in data[primary_key_value].items():
                    worksheet.cell(row=i, column=col_ind).value = val

        wb.save(output_file_path)


if __name__ == "__main__":
    a = PyExcel.get_data_simple(
        'resources/待爬虫学校名单.xlsx', '待爬虫链接', {'类型': 1, '名称': 2, '链接': 3})
    # print(a)

    import sys
    import os
    CURRENT_DIR = os.path.split(os.path.abspath(__file__))[0]  # 当前目录
    config_path = CURRENT_DIR.rsplit('\\', 1)[0]  # 上二级目录
    if config_path not in sys.path:
        sys.path.append(config_path)
    from classes import School

    school_infos = [
        School({'floor': '小学', 'name': '番禺区京师奥园南奥实验', 'url': 'https://www.sxkid.com/school-hlxgj/', 'baidu_urls': [
               'http://www.baidu.com/link?url=FdlMA6DPfVhKZ0WYfPhQShpWDxmGbj53_GvWlAuF8jw4xuU4DiLmtE-jXGbj6LCGY5YwuZObwcFUKlhzgkExMzvJ3v78MjTsqiJBSItsWFqUePGLvdZQoE4NevE11hFb'], 'is_baidu_collected': True, 'no': 636941, 'uuid': '108c8099-96a4-461f-b435-f27e6809c8b5'}),
        School({'floor': '小学', 'name': '华美实验', 'url': 'https://www.sxkid.com/school-fylfp/', 'baidu_urls': [
               'http://www.baidu.com/link?url=FlbKw-xof_l1xACpGvAVaPKU_UDmfk1SyUpoIvnkFghC4NPsb2LLgBk5wG_vLTYtqeZwfoMWRHSZaCf2c32ZaLOfBGYgeQtwMaPsBltR9DLIKIF1yX4ufdW-tacGR4ygfSRDvAVfnERjr4HOP1vzMq'], 'is_baidu_collected': True, 'no': 355719, 'uuid': '484cb2a1-1db2-4cf8-b869-875e704b221e'})
    ]
    data = {
        school_info.url: {
            5: str(school_info.is_baidu_collected)
        }
        for school_info in school_infos
    }

    PyExcel.write_data(
        'resources/待爬虫学校名单.xlsx',
        'resources/待爬虫学校名单v2.xlsx',
        '待爬虫链接',
        4,
        data
    )
